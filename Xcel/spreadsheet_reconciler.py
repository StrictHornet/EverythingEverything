from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(
    filename="C:/Users/Ehi/Google Drive/Work/VODACOM MAIN COPY OF SURVEY SPREADSHEET.xlsx")

################################
# EVERYTHING IN ONE FILE

# For loop for search and cell assignment
for ir in workbook["RecNew"].iter_rows(min_row=2, min_col=3):
    # print(ir) # For debugging
    for sheet in workbook:
        for row in sheet.iter_rows(min_row=2, min_col=8):
            # print(row) # For debugging
            try:
                if ir[0].value == row[0].value:
                    if row[6].value != "NIL":
                        # print("Present") # For debugging
                        # print("VBN is {} and Recon is {}".format(
                        #     row[0].value, ir[0].value))
                        ir[21].value = "Failed"  # 19
                        ir[22].value = row[6].value  # 20
                        break  # Since IR has been found loop should break to next IR
                    else:
                        ir[21].value = "OK"
            except:
                print("Didn't work!")
        else:
            continue
        break  # Break the outer loop

# For loop for duplicate search
for ir in workbook["RecNew"].iter_rows(min_row=2, min_col=3):
    for sheet in workbook:
        for row in sheet.iter_rows(min_row=2, min_col=3):
            try:
                if ir[0].value == row[0].value:
                    if row[21].value == "Approved":
                        # print("Duplicate") # For debugging
                        # print("VBN is {} and Recon is {}".format(
                        #     row[0].value, ir[0].value))
                        ir[21].value = "DUPLICATE"  # 19
                        ir[22].value = row[22].value  # 20
                        break  # Since IR has been found loop should break to next IR
            except:
                print("Didn't work!")
        else:
            continue
        break  # Break the outer loop

workbook.save(
    filename="C:/Users/Ehi/Google Drive/Work/VODACOM MAIN COPY OF SURVEY SPREADSHEET.xlsx")
